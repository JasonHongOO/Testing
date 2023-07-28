from openpyxl import load_workbook
import pandas as pd
import os
import json


def Creat(data):
    # wb = openpyxl.Workbook()           # 建立空白的 Excel 活頁簿物件
    # sheet = wb.create_sheet('mysheet')     # 建立空白的工作表

    # data = [["DisplayName","Language","PictureUrl","StatusMessage","userId"]]   # 二維陣列資料
    # for i in data:
    #     sheet.append(i)

    # wb.save('DataBase.xlsx')


    # print(sheet.title, sheet.max_row, sheet.max_column)
    print("ok")

def SaveData(Profile, file_name='DataBase.xlsx'):
    #轉換成 dictionary
    data = {"DisplayName": Profile.display_name, "UserId": Profile.user_id, "Language": Profile.language, "StatusMessage": Profile.status_message, "PictureUrl": Profile.picture_url}

    # 檢查是否存在 Excel 檔案，如果不存在，創建新的 DataFrame 並寫入標題
    if not os.path.isfile(file_name):
        columns = ["DisplayName", "UserId", "Language",  "StatusMessage", "PictureUrl"]
        df = pd.DataFrame(columns=columns)
        df.to_excel(file_name, index=False)
    
    # 將資料加入 DataFrame
    df = pd.read_excel(file_name)
    if 'UserId' in df.columns and data['UserId'] in df['UserId'].values:
        print("新用戶加入 : 資料已在資料庫中，不再增新。")
    else:
        df = pd.concat([df, pd.DataFrame(data, index=[0])], ignore_index=True)
        print("新用戶加入 : 資料不在資料庫中，已增新完畢。")

    # 將 DataFrame 寫回 Excel 檔案
    df.to_excel(file_name, index=False)


    wb = load_workbook(file_name, data_only=True)
    s1 = wb.active

    for letter in range(ord('A'), ord('E')+1):
        # print(chr(letter))
        s1.column_dimensions[chr(letter)].width = 28

    wb.save(file_name)


# s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
# s2 = wb.active           # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )


if __name__ == "__main__":
    data = {"DisplayName": "John", "Language": "English", "PictureUrl": "https://example.com/john.png", "StatusMessage": "Hello", "userId": "user123"},

    SaveData(data, file_name='DataBase.xlsx')